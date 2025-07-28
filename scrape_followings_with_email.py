# scraper_no_proxy.py
# ------------------
# Instagram followings scraper with:
# - Single-session (no proxies)
# - Incremental CSV write on every profile (saves progress if stopped)
# - Scrapes followers count for each profile
# - Rate-limit backoff, spam-feedback handling, and random delays
# - Exponential backoff on 400 feedback_required, 429, 503
# - Hourly cool-down to reduce ban risk
# - Dynamic long pauses every 120-150 profiles
# - Deduplication of profiles

import requests
import time
import csv
import random
import pandas as pd
from urllib.parse import urlparse
from openpyxl import load_workbook
import os
import re
from datetime import datetime, timedelta

# 🚀 Instagram session cookies (rotate through multiple accounts to avoid blocks)
from itertools import cycle

ACCOUNTS = [
    # Account 1
    {
        "sessionid": "75787062789%3AHpFaBGBTcBaRer%3A25%3AAYdsi2EjjAzaPpQ8fuMZkwy_nfQmTgaQ7JodYC-qpw",
        "ds_user_id": "75787062789",
        "csrftoken": "MRHrdFkgeom6BRYiCe7G3oR1X7ZTuN0r"
    },
    # Account 2
    {
        "sessionid": "76155323513%3A7aqrgyX3aApBu0%3A16%3AAYdjP8772HdoYEF9_EQjZcITL8RE2jQKDuaH94Qkwg",
        "ds_user_id": "76155323513",
        "csrftoken": "hzzfuoacMKr08b8oZnUtUsVMepz7bHwk"
    },
    # Account 3 (your latest)
    {
        "sessionid": "75507198098%3AlNClvBV0OOJTmo%3A9%3AAYdal7i6vhArtnfiLz6hQyHe8OicU8ButJcgCw6LRg",
        "ds_user_id": "75507198098",
        "csrftoken": "TgznZOk-izLYFDnY4yx6J_"
    }
]
# cycle for rotating sessions
det_account_cycle = cycle(ACCOUNTS)


# Common request headers
HEADERS = {
    "User-Agent": "Instagram 219.0.0.12.117 Android",
    "Accept": "*/*",
    "Accept-Language": "en-US",
    "X-IG-App-ID": "936619743392459",
    "Referer": "https://www.instagram.com/",
    "X-Requested-With": "XMLHttpRequest"
}

SEED_CSV = "seed_accounts.csv"
OUTPUT_XLSX = "followings_output.xlsx"
OUTPUT_CSV = "followings_progress.csv"

# Prepare incremental CSV writer
import sys  # for exit on permission errors
csv_file_exists = os.path.exists(OUTPUT_CSV)
try:
    csv_file = open(OUTPUT_CSV, 'a', newline='', encoding='utf-8')
except PermissionError:
    print(f"❌ Permission denied: please close '{OUTPUT_CSV}' if it's open in another program and try again.")
    sys.exit(1)

csv_fieldnames = [
    'seed_account', 'username', 'profile_url', 'full_name',
    'bio', 'email_from_button', 'email_from_bio', 'followers_count'
]
csv_writer = csv.DictWriter(csv_file, fieldnames=csv_fieldnames)
if not csv_file_exists:
    csv_writer.writeheader()

# Deduplication set loaded from previous progress
seen_usernames = set()
if os.path.exists(OUTPUT_CSV):
    with open(OUTPUT_CSV, newline='', encoding='utf-8') as f2:
        reader2 = csv.DictReader(f2)
        for r2 in reader2:
            seen_usernames.add(r2['username'])
    print(f"ℹ️ Loaded {len(seen_usernames)} previously scraped usernames from {OUTPUT_CSV}")
else:
    print("ℹ️ No previous progress file found, starting fresh.")

# Track runtime for hourly cooldown
start_time = datetime.now()


def get_session():
    sess = requests.Session()
    sess.cookies.update(ACCOUNTS[0])
    sess.headers.update(HEADERS)
    return sess


def make_request(url, params=None, max_retries=5):
    backoff = 5
    for attempt in range(max_retries):
        sess = get_session()
        try:
            res = sess.get(url, params=params, timeout=15)
        except requests.RequestException as e:
            print(f"⚠️ Req error: {e}. Backoff {backoff}s...")
            time.sleep(backoff + random.uniform(0,2))
            backoff = min(backoff * 2, 120)
            continue
        # handle spam feedback
        if res.status_code == 400:
            try:
                data = res.json()
                if data.get('message') == 'feedback_required':
                    print(f"⏱️ Feedback required. Backoff {backoff}s...")
                    time.sleep(backoff + random.uniform(0,2))
                    backoff = min(backoff * 2, 360)
                    continue
            except ValueError:
                pass
        # handle rate limits
        if res.status_code in (429, 503):
            print(f"⏱️ Rate limited {res.status_code}. Backoff {backoff}s...")
            time.sleep(backoff + random.uniform(0,2))
            backoff = min(backoff * 2, 360)
            continue
        if res.status_code == 200:
            return res
        print(f"⚠️ HTTP {res.status_code}. Response: {res.text[:100]}")
        return None
    print(f"❌ All retries failed for URL: {url}")
    return None


def extract_username(url):
    return urlparse(url).path.strip('/').split('/')[0]


def get_user_id(username):
    api = f"https://i.instagram.com/api/v1/users/web_profile_info/?username={username}"
    res = make_request(api)
    if res:
        return res.json().get('data', {}).get('user', {}).get('id')
    return None


def get_profile_details(user_id):
    api = f"https://i.instagram.com/api/v1/users/{user_id}/info/"
    res = make_request(api)
    if not res:
        return {}
    u = res.json().get('user', {})
    return {
        'full_name': u.get('full_name',''),
        'bio': u.get('biography',''),
        'public_email': u.get('public_email',''),
        'followers_count': u.get('follower_count', 0)
    }


def extract_email_from_bio(bio):
    m = re.search(r"[A-Za-z0-9_.+-]+@[A-Za-z0-9-]+\.[A-Za-z0-9-.]+", bio)
    return m.group(0) if m else ''


def scrape_followings(user_id, seed_name):
    global start_time
    after = ''
    has_next = True
    results = []
    count = 0
    # Random threshold for longer pauses
    next_pause = random.randint(120, 150)

    while has_next:
        # hourly cooldown
        if datetime.now() - start_time > timedelta(hours=1):
            print("⏰ Hourly break: sleeping 5-6 minutes...")
            time.sleep(300 + random.uniform(0,60))
            start_time = datetime.now()

        url = f"https://i.instagram.com/api/v1/friendships/{user_id}/following/"
        params = {'count': 50, 'max_id': after} if after else {'count': 50}
        res = make_request(url, params)
        if not res:
            break
        data = res.json()

        for u in data.get('users', []):
            uname = u['username']
            if uname in seen_usernames:
                continue
            seen_usernames.add(uname)

            details = get_profile_details(u['pk'])
            email_bio = extract_email_from_bio(details.get('bio',''))
            row = {
                'seed_account': seed_name,
                'username': uname,
                'profile_url': f"https://instagram.com/{uname}",
                'full_name': details.get('full_name',''),
                'bio': details.get('bio',''),
                'email_from_button': details.get('public_email',''),
                'email_from_bio': email_bio,
                'followers_count': details.get('followers_count', 0)
            }
            # record immediately
            results.append(row)
            csv_writer.writerow(row)
            csv_file.flush()

            count += 1
            time.sleep(random.uniform(2.5, 4.5))

            if count >= next_pause:
                pause = random.uniform(25, 50)
                print(f"⏸️ Pausing after {count} profiles for {pause:.0f}s...")
                time.sleep(pause)
                next_pause = count + random.randint(120, 150)

        after = data.get('next_max_id')
        has_next = bool(after)
        time.sleep(random.uniform(2.0, 4.0))

    return results


def main():
    with open(SEED_CSV, newline='', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        seeds = [(r['seed_account'].strip(), extract_username(r['profile_link'])) for r in reader if r.get('profile_link')]

    if os.path.exists(OUTPUT_XLSX):
        writer = pd.ExcelWriter(OUTPUT_XLSX, engine='openpyxl', mode='a', if_sheet_exists='overlay')
        existing = load_workbook(OUTPUT_XLSX).sheetnames
    else:
        writer = pd.ExcelWriter(OUTPUT_XLSX, engine='openpyxl', mode='w')
        existing = []

    for seed_name, uname in seeds:
        sheet = seed_name[:30]
        if sheet in existing:
            print(f"⏭️ Skipping {seed_name}: already exists.")
            continue
        print(f"🔍 Scraping @{uname} (seed: {seed_name})")
        uid = get_user_id(uname)
        if not uid:
            print(f"❌ Could not get ID for @{uname}")
            continue

        data = scrape_followings(uid, seed_name)
        if data:
            df = pd.DataFrame(data)
            df.to_excel(writer, sheet_name=sheet, index=False)
            print(f"✅ Saved {len(data)} profiles from {seed_name}")
        else:
            print(f"⚠️ No data for {seed_name}")

        time.sleep(random.uniform(10, 20))

    writer.close()
    csv_file.close()
    print("\n📁 Done! Check:", OUTPUT_XLSX, "and CSV:", OUTPUT_CSV)

if __name__ == '__main__':
    main()
